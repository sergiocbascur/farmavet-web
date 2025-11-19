#!/usr/bin/env python3
"""
Script para importar metodologías desde un archivo Excel
Uso: python importar_metodologias_excel.py "RESUMEN CLIENTES-LAB.xlsx"
"""

import sys
import os
import sqlite3
from datetime import datetime

# Intentar importar pandas u openpyxl
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    try:
        import openpyxl
        HAS_OPENPYXL = True
        HAS_PANDAS = False
    except ImportError:
        print("[ERROR] Necesitas instalar pandas o openpyxl")
        print("   Instala con: pip install pandas openpyxl")
        sys.exit(1)

def get_db_path():
    """Obtiene la ruta de la base de datos"""
    # Buscar en varios lugares posibles
    possible_paths = [
        'instance/database.db',  # Prioridad 1: estructura Flask estándar
        'farmavet_web.db',
        'database.db'
    ]
    
    for path in possible_paths:
        if os.path.exists(path):
            return path
    
    # Si no se encuentra, preguntar
    print("[AVISO] No se encontro la base de datos automaticamente.")
    db_path = input("Ingresa la ruta a la base de datos: ").strip()
    if os.path.exists(db_path):
        return db_path
    else:
        print(f"[ERROR] No se encontro el archivo: {db_path}")
        sys.exit(1)

def normalize_text(text):
    """Normaliza texto para búsqueda"""
    if not text or pd.isna(text):
        return None
    return str(text).strip()

def import_from_excel(excel_path, db_path, dry_run=False):
    """Importa metodologías desde Excel a la base de datos"""
    
    if not os.path.exists(excel_path):
        print(f"[ERROR] No se encontro el archivo: {excel_path}")
        sys.exit(1)
    
    print(f"Leyendo archivo Excel: {excel_path}")
    
    # Leer Excel - Estructura: ITLF (método), MÉTODO, Analito, EQUIPO, LD, LC, MATRIZ, ACREDITADO
    # Los encabezados están en la fila 2 (índice 2)
    # Los datos empiezan en la fila 4 (índice 4)
    try:
        if HAS_PANDAS:
            # Leer sin header para procesar manualmente
            df_raw = pd.read_excel(excel_path, engine='openpyxl', header=None)
            
            # Los datos empiezan en la fila 4 (índice 4)
            data_rows = []
            current_itlf = None
            current_metodo = None
            current_equipo = None
            
            for i in range(4, len(df_raw)):  # Empezar desde fila 4
                row = df_raw.iloc[i]
                
                # Leer valores según estructura:
                # Col 1: ITLF, Col 2: MÉTODO, Col 3: Analito, Col 4: EQUIPO, Col 5: LD, Col 6: LC, Col 7: MATRIZ, Col 8: ACREDITADO
                itlf = row[1] if pd.notna(row[1]) else None
                metodo = row[2] if pd.notna(row[2]) else None
                analito = row[3] if pd.notna(row[3]) else None
                equipo = row[4] if pd.notna(row[4]) else None
                ld = row[5] if pd.notna(row[5]) else None
                lc = row[6] if pd.notna(row[6]) else None
                matriz = row[7] if pd.notna(row[7]) else None
                acreditado = row[8] if pd.notna(row[8]) else None
                
                # Si hay ITLF, actualizar método actual
                if pd.notna(itlf) and str(itlf).strip():
                    current_itlf = str(itlf).strip()
                    current_metodo = str(metodo).strip() if pd.notna(metodo) else None
                    current_equipo = str(equipo).strip() if pd.notna(equipo) else None
                
                # Si hay analito, crear una entrada
                if pd.notna(analito) and str(analito).strip() and current_itlf:
                    data_rows.append({
                        'itlf': current_itlf,
                        'metodo': current_metodo or str(analito).strip(),
                        'analito': str(analito).strip(),
                        'equipo': current_equipo or str(equipo).strip() if pd.notna(equipo) else None,
                        'ld': str(ld).strip() if pd.notna(ld) else None,
                        'lc': str(lc).strip() if pd.notna(lc) else None,
                        'matriz': str(matriz).strip() if pd.notna(matriz) else None,
                        'acreditado': str(acreditado).strip() if pd.notna(acreditado) else None,
                    })
            
            df = pd.DataFrame(data_rows)
            
        else:
            # Usar openpyxl directamente
            from openpyxl import load_workbook
            wb = load_workbook(excel_path)
            ws = wb.active
            
            data_rows = []
            current_itlf = None
            current_metodo = None
            current_equipo = None
            
            for row in ws.iter_rows(min_row=4, values_only=True):
                itlf = row[1] if row[1] else None
                metodo = row[2] if row[2] else None
                analito = row[3] if row[3] else None
                equipo = row[4] if row[4] else None
                ld = row[5] if row[5] else None
                lc = row[6] if row[6] else None
                matriz = row[7] if row[7] else None
                acreditado = row[8] if row[8] else None
                
                if itlf:
                    current_itlf = str(itlf).strip()
                    current_metodo = str(metodo).strip() if metodo else None
                    current_equipo = str(equipo).strip() if equipo else None
                
                if analito and current_itlf:
                    data_rows.append({
                        'itlf': current_itlf,
                        'metodo': current_metodo or str(analito).strip(),
                        'analito': str(analito).strip(),
                        'equipo': current_equipo or str(equipo).strip() if equipo else None,
                        'ld': str(ld).strip() if ld else None,
                        'lc': str(lc).strip() if lc else None,
                        'matriz': str(matriz).strip() if matriz else None,
                        'acreditado': str(acreditado).strip() if acreditado else None,
                    })
            
            df = pd.DataFrame(data_rows)
        
        print(f"[OK] Archivo leido: {len(df)} analitos encontrados")
        print(f"[INFO] Metodos unicos (ITLF): {df['itlf'].nunique() if 'itlf' in df.columns else 0}")
        
    except Exception as e:
        print(f"[ERROR] Error leyendo Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    
    # Conectar a la base de datos
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Verificar que la tabla existe
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='metodologias'")
    if not cursor.fetchone():
        print("[ERROR] La tabla 'metodologias' no existe en la base de datos")
        conn.close()
        sys.exit(1)
    
    # Las columnas ya están mapeadas en el DataFrame
    # Estructura: itlf, metodo, analito, equipo, ld, lc, matriz, acreditado
    print(f"\n[INFO] Estructura detectada:")
    print(f"   - ITLF (código método): Columna 'itlf'")
    print(f"   - MÉTODO (nombre): Columna 'metodo'")
    print(f"   - Analito: Columna 'analito'")
    print(f"   - EQUIPO (técnica): Columna 'equipo'")
    print(f"   - LD: Columna 'ld'")
    print(f"   - LC: Columna 'lc'")
    print(f"   - MATRIZ: Columna 'matriz'")
    print(f"   - ACREDITADO: Columna 'acreditado'")
    
    # Verificar que las columnas necesarias existen
    required_columns = ['analito']
    missing_columns = [col for col in required_columns if col not in df.columns]
    
    if missing_columns:
        print(f"\n[ERROR] Faltan columnas requeridas: {', '.join(missing_columns)}")
        print(f"   Columnas disponibles: {', '.join(df.columns.tolist())}")
        conn.close()
        sys.exit(1)
    
    # Categoría por defecto
    categoria_default = 'residuos'  # Se puede ajustar según necesites
    
    # Procesar cada fila
    imported = 0
    skipped = 0
    errors = []
    
    print(f"\nProcesando {len(df)} filas...")
    
    for idx, row in df.iterrows():
        try:
            # Extraer valores directamente del DataFrame ya procesado
            itlf = normalize_text(row.get('itlf'))
            metodo = normalize_text(row.get('metodo'))
            analito = normalize_text(row.get('analito'))
            equipo = normalize_text(row.get('equipo'))
            ld = normalize_text(row.get('ld'))
            lc = normalize_text(row.get('lc'))
            matriz = normalize_text(row.get('matriz'))
            acreditado_val = normalize_text(row.get('acreditado'))
            
            # Si no hay analito, saltar esta fila
            if not analito:
                skipped += 1
                continue
            
            # Usar ITLF como código, método como nombre
            codigo = itlf if itlf else None
            nombre = metodo if metodo else analito
            
            # Mapear abreviaciones de matriz a nombres completos
            matriz_mapping = {
                'pp': 'Productos pecuarios',
                'phb': 'Productos hidrobiológicos',
                'PHB': 'Productos hidrobiológicos',
                'PP': 'Productos pecuarios'
            }
            
            # Aplicar mapeo si existe
            if matriz and matriz.lower() in matriz_mapping:
                matriz = matriz_mapping[matriz.lower()]
            elif matriz:
                # Normalizar: primera letra mayúscula, resto minúsculas
                matriz = matriz.strip()
            
            # Si no hay matriz, usar un valor por defecto
            if not matriz:
                matriz = 'No especificada'
            
            # Categoría (por defecto)
            categoria = categoria_default
            
            # Técnica (EQUIPO)
            tecnica = equipo if equipo else None
            
            # Límites
            limite_deteccion = ld if ld else None
            limite_cuantificacion = lc if lc else None
            
            # Acreditada (puede ser "OK", "Sí", etc.)
            acreditada = False
            if acreditado_val:
                acreditado_str = str(acreditado_val).lower().strip()
                acreditada = acreditado_str in ['ok', 'si', 'sí', 's', 'yes', 'true', '1', 'acreditada', 'acreditado']
            
            # Otros campos opcionales
            nombre_en = None
            analito_en = None
            matriz_en = None
            tecnica_en = None
            norma_referencia = None
            vigencia = None
            orden = None
            
            # Validar campos requeridos
            if not analito:
                skipped += 1
                errors.append(f"Fila {idx+2}: Falta campo requerido (analito/método)")
                continue
            
            # Verificar si ya existe (por código+analito+matriz para permitir múltiples analitos por método)
            # IMPORTANTE: Buscar por código+analito+matriz para que cada analito sea una fila separada
            if codigo:
                cursor.execute("SELECT id FROM metodologias WHERE codigo = ? AND analito = ? AND matriz = ?", 
                             (codigo, analito, matriz))
            else:
                cursor.execute("SELECT id FROM metodologias WHERE nombre = ? AND analito = ? AND matriz = ?", 
                             (nombre, analito, matriz))
            
            existing = cursor.fetchone()
            
            if existing:
                if not dry_run:
                    # Actualizar existente
                    cursor.execute('''
                        UPDATE metodologias 
                        SET codigo=?, nombre=?, nombre_en=?, categoria=?, analito=?, analito_en=?,
                            matriz=?, matriz_en=?, tecnica=?, tecnica_en=?, limite_deteccion=?,
                            limite_cuantificacion=?, norma_referencia=?, vigencia=?, acreditada=?, orden=?
                        WHERE id=?
                    ''', (codigo, nombre, nombre_en, categoria, analito, analito_en, matriz, matriz_en,
                          tecnica, tecnica_en, limite_deteccion, limite_cuantificacion, norma_referencia,
                          vigencia, acreditada, orden, existing['id']))
                    print(f"   [ACTUALIZADO] {nombre} - {analito} en {matriz}")
                else:
                    print(f"   [DRY RUN] Actualizaria: {nombre} - {analito} en {matriz}")
            else:
                if not dry_run:
                    # Insertar nuevo
                    cursor.execute('''
                        INSERT INTO metodologias 
                        (codigo, nombre, nombre_en, categoria, analito, analito_en, matriz, matriz_en,
                         tecnica, tecnica_en, limite_deteccion, limite_cuantificacion, norma_referencia,
                         vigencia, acreditada, orden, activo)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
                    ''', (codigo, nombre, nombre_en, categoria, analito, analito_en, matriz, matriz_en,
                          tecnica, tecnica_en, limite_deteccion, limite_cuantificacion, norma_referencia,
                          vigencia, acreditada, orden))
                    print(f"   [INSERTADO] {nombre} - {analito} en {matriz}")
                else:
                    print(f"   [DRY RUN] Insertaria: {nombre} - {analito} en {matriz}")
            
            imported += 1
            
        except Exception as e:
            skipped += 1
            errors.append(f"Fila {idx+2}: {str(e)}")
            print(f"   [ERROR] Error en fila {idx+2}: {str(e)}")
    
    if not dry_run:
        conn.commit()
        print(f"\n[OK] Importacion completada!")
    else:
        print(f"\n[DRY RUN] Simulacion completada (no se modifico la base de datos)")
    
    print(f"   [OK] Importadas/Actualizadas: {imported}")
    print(f"   [AVISO] Omitidas: {skipped}")
    
    if errors:
        print(f"\n[AVISO] Errores encontrados ({len(errors)}):")
        for error in errors[:10]:  # Mostrar solo los primeros 10
            print(f"   - {error}")
        if len(errors) > 10:
            print(f"   ... y {len(errors) - 10} errores más")
    
    conn.close()
    return imported, skipped, len(errors)

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Uso: python importar_metodologias_excel.py <archivo_excel> [--dry-run] [--yes]")
        print("\nEjemplo:")
        print("  python importar_metodologias_excel.py 'RESUMEN CLIENTES-LAB.xlsx'")
        print("  python importar_metodologias_excel.py 'RESUMEN CLIENTES-LAB.xlsx' --dry-run")
        print("  python importar_metodologias_excel.py '/ruta/completa/al/archivo.xlsx'")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    dry_run = '--dry-run' in sys.argv or '-n' in sys.argv
    skip_confirm = '--yes' in sys.argv or '-y' in sys.argv or dry_run
    
    # Si el archivo no existe, intentar buscar en ubicaciones comunes
    if not os.path.exists(excel_path):
        print(f"⚠️  No se encontró el archivo: {excel_path}")
        print("\nBuscando en ubicaciones comunes...")
        
        # Buscar en el directorio actual y subdirectorios
        posibles_nombres = [
            'RESUMEN CLIENTES-LAB.xlsx',
            'resumen clientes-lab.xlsx',
            'RESUMEN_CLIENTES_LAB.xlsx',
            'resumen_clientes_lab.xlsx'
        ]
        
        encontrado = False
        for nombre in posibles_nombres:
            # Buscar en directorio actual
            if os.path.exists(nombre):
                excel_path = nombre
                encontrado = True
                print(f"✅ Encontrado: {os.path.abspath(excel_path)}")
                break
            
            # Buscar en directorio padre
            parent_path = os.path.join('..', nombre)
            if os.path.exists(parent_path):
                excel_path = parent_path
                encontrado = True
                print(f"✅ Encontrado: {os.path.abspath(excel_path)}")
                break
        
        if not encontrado:
            print("\n❌ No se encontró el archivo Excel en ubicaciones comunes.")
            print("\nOpciones:")
            print("  1. Especifica la ruta completa del archivo:")
            print("     python importar_metodologias_excel.py '/ruta/completa/RESUMEN CLIENTES-LAB.xlsx'")
            print("  2. Copia el archivo al directorio actual del script")
            print(f"     Directorio actual: {os.getcwd()}")
            sys.exit(1)
    
    if dry_run:
        print("[DRY RUN] No se modificara la base de datos\n")
    
    db_path = get_db_path()
    
    print(f"Base de datos: {db_path}")
    print(f"Archivo Excel: {excel_path}\n")
    
    if not skip_confirm:
        confirm = input("Continuar con la importacion? (s/n): ").strip().lower()
        if confirm not in ['s', 'si', 'sí', 'y', 'yes']:
            print("[CANCELADO] Importacion cancelada")
            sys.exit(0)
    
    import_from_excel(excel_path, db_path, dry_run=dry_run)

